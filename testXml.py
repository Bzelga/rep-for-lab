from openpyxl import load_workbook
import ftplib

HOSTNAME = "46.185.210.235"
USERNAME = "bzelga"
PASSWORD = "qwer"
ftp_server = ftplib.FTP(HOSTNAME, USERNAME, PASSWORD)
ftp_server.retrbinary('testxlsx.xlsx', file)

# wb = load_workbook('C://Users//Bzelga Notebook//Desktop//20102021_1_TJD.xlsx')
wb = load_workbook('')

ws = wb.get_sheet_by_name('Лист1')

result = {'num_dog':[], 'date_usl':[],'gruz':[],'stan_otpr':[],'stan_nazn':[],'date_otpr':[],'num_nakl':[],'kol_vag':[],'kol_ton':[],'cen_ton':[],'sumbeznds':[],'sumnds':[], \
    'sumsnds':[], 'numact':[]}

for i in range(6, ws.max_row):
    result['num_dog'].append(ws.cell(row = i, column = 1).value)
    result['date_usl'].append(ws.cell(row = i, column = 2).value)
    result['gruz'].append(ws.cell(row = i, column = 3).value)
    result['stan_otpr'].append(ws.cell(row = i, column = 4).value)
    result['stan_nazn'].append(ws.cell(row = i, column = 5).value)
    result['date_otpr'].append(ws.cell(row = i, column = 6).value)
    result['num_nakl'].append(ws.cell(row = i, column = 7).value)
    result['kol_vag'].append(ws.cell(row = i, column = 8).value)
    result['kol_ton'].append(ws.cell(row = i, column = 9).value)
    result['cen_ton'].append(ws.cell(row = i, column = 10).value)
    result['sumbeznds'].append(ws.cell(row = i, column = 11).value)
    result['sumnds'].append(ws.cell(row = i, column = 12).value)
    result['sumsnds'].append(ws.cell(row = i, column = 13).value)
    result['numact'].append(ws.cell(row = i, column = 14).value)

print(result)